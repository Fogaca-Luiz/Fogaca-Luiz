"""
Atualiza a planilha Itau.xlsx com os lancamentos reais encontrados nos
extratos .xls salvos na subpasta "Itau".

Uso:
    python atualizar_extrato_itau.py

Regras aplicadas:
  - So sao considerados lancamentos reais (linhas "SALDO ANTERIOR" e
    "SALDO TOTAL DISPONIVEL DIA" sao ignoradas, pois sao marcadores de
    saldo). A secao final "lancamentos futuros / saidas futuras" de cada
    extrato tambem e ignorada (lancamentos ainda nao efetivados).
  - Duplicatas entre extratos consecutivos que se sobrepoem (mesma data +
    mesmo texto + mesmo valor) sao descartadas, mas a contagem de
    ocorrencias e respeitada -- dois lancamentos genuinamente iguais no
    mesmo dia (mesmo texto e valor) sao mantidos.
  - O periodo coberto pelos extratos e definido pela primeira e ultima
    data encontradas nos arquivos .xls da pasta Itau. Dentro desse
    periodo, a planilha e comparada DIA A DIA com o extrato:
      - Se os valores de um dia na planilha ja batem exatamente com os
        valores do extrato para aquele dia (mesmo conjunto de creditos e
        debitos), a linha e mantida como esta -- isso preserva historico
        ja reconciliado (mesmo que os extratos, por causa do "SALDO
        ANTERIOR" do Itau, repitam 1-2 dias que ja estavam corretos).
        Nesses dias, se a coluna LANCAMENTO estiver em branco, ela e
        preenchida cruzando por valor com o texto bruto do extrato; e se
        DESCRICAO e/ou CATEGORIA estiverem em branco mas o mesmo texto de
        LANCAMENTO ja tiver sido classificado em outra linha da planilha,
        essa classificacao e reaplicada retroativamente ali tambem.
      - Se os valores nao baterem (dia orcado/planejado ou dia ausente),
        as linhas existentes desse dia sao removidas e substituidas
        pelos lancamentos reais do extrato.
    Linhas com data fora desse periodo (historico antigo ou orcamento
    futuro ainda nao coberto por um extrato) nunca sao tocadas.
  - Nas linhas novas, a coluna LANCAMENTO recebe o texto bruto do
    extrato. DESCRICAO e CATEGORIA sao preenchidas automaticamente quando
    o mesmo texto de LANCAMENTO ja apareceu antes em alguma linha da
    planilha (incluindo as que estao para ser substituidas) que tenha
    essas colunas preenchidas -- esse "aprendizado" e capturado antes de
    remover qualquer linha, para nao se perder. Quando o texto e inedito,
    as duas colunas ficam em branco para preenchimento manual.
"""

import datetime
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path

import pandas as pd
import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR / "Itau.xlsx"
XLS_DIR = SCRIPT_DIR / "Itau"
SHEET_NAME = "Itau"

COL_DATA = 1
COL_LANCAMENTO = 2
COL_DESCRICAO = 3
COL_CREDITO = 4
COL_DEBITO = 5
COL_SALDO = 6
COL_CATEGORIA = 7

SALDO_MARKERS = ("SALDO ANTERIOR", "SALDO TOTAL")


def parse_data(valor):
    if not isinstance(valor, str):
        return None
    try:
        return datetime.datetime.strptime(valor.strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def extrair_lancamentos(caminho_xls):
    """Le um extrato .xls e retorna uma lista de (data, texto, valor)."""
    df = pd.read_excel(caminho_xls, header=None)
    registros = []
    dentro_da_tabela = False

    for _, linha in df.iterrows():
        data = parse_data(linha[0])

        if data is None:
            if dentro_da_tabela:
                # Fim do bloco de lancamentos reais (comeca a secao de
                # "lancamentos futuros" ou fim do arquivo).
                break
            continue

        dentro_da_tabela = True
        texto = linha[1]
        valor = linha[3]

        if not isinstance(texto, str):
            continue
        texto = texto.strip()

        if any(texto.upper().startswith(m) for m in SALDO_MARKERS):
            continue

        if pd.isna(valor):
            continue

        registros.append((data, texto, float(valor)))

    return registros


def coletar_transacoes(arquivos):
    """Le todos os extratos e devolve uma lista deduplicada e ordenada
    cronologicamente de (data, texto, valor)."""
    lancamentos_por_arquivo = []
    for arquivo in arquivos:
        registros = extrair_lancamentos(arquivo)
        if registros:
            lancamentos_por_arquivo.append((registros[0][0], arquivo.name, registros))
    lancamentos_por_arquivo.sort(key=lambda item: item[0])

    ja_vistos = Counter()
    transacoes = []
    for _, nome_arquivo, registros in lancamentos_por_arquivo:
        adicionados_neste_arquivo = 0
        ocorrencias_no_arquivo = Counter()
        for data, texto, valor in registros:
            chave = (data, texto, valor)
            ocorrencias_no_arquivo[chave] += 1
            if ja_vistos[chave] >= ocorrencias_no_arquivo[chave]:
                continue
            ja_vistos[chave] += 1
            transacoes.append(chave)
            adicionados_neste_arquivo += 1
        print(f"{nome_arquivo}: {adicionados_neste_arquivo} lancamento(s) real(is)")

    transacoes.sort(key=lambda item: item[0])
    return transacoes


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {XLSX_PATH}")
    if not XLS_DIR.exists():
        raise FileNotFoundError(f"Pasta de extratos nao encontrada: {XLS_DIR}")

    arquivos = sorted(XLS_DIR.glob("*.xls"))
    if not arquivos:
        print("Nenhum arquivo .xls encontrado em", XLS_DIR)
        return

    transacoes = coletar_transacoes(arquivos)
    if not transacoes:
        print("Nenhum lancamento real encontrado nos extratos.")
        return

    data_inicio = transacoes[0][0]
    data_fim = transacoes[-1][0]
    print(f"Periodo coberto pelos extratos: {data_inicio:%d/%m/%Y} a {data_fim:%d/%m/%Y}")

    extrato_por_dia = defaultdict(list)
    for data, texto, valor in transacoes:
        extrato_por_dia[data].append((texto, valor))

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb[SHEET_NAME]
    ultima_linha_planilha = ws.max_row

    # Aprende, a partir de TODA a planilha (inclusive linhas que podem
    # vir a ser substituidas), qual DESCRICAO/CATEGORIA foi usada para
    # cada texto de LANCAMENTO -- para reaplicar automaticamente. Isso e
    # capturado antes de remover qualquer linha.
    mapa_desc_cat = {}
    for r in range(2, ultima_linha_planilha + 1):
        texto_existente = ws.cell(row=r, column=COL_LANCAMENTO).value
        if not texto_existente:
            continue
        texto_existente = texto_existente.strip()
        descricao_existente = ws.cell(row=r, column=COL_DESCRICAO).value
        categoria_existente = ws.cell(row=r, column=COL_CATEGORIA).value
        if descricao_existente or categoria_existente:
            mapa_desc_cat[texto_existente] = (descricao_existente, categoria_existente)

    # Linhas ja existentes cuja data cai dentro do periodo coberto pelos
    # extratos.
    linhas_candidatas = []
    primeira_linha_apos_periodo = None
    for r in range(2, ultima_linha_planilha + 1):
        data_linha = ws.cell(row=r, column=COL_DATA).value
        if isinstance(data_linha, datetime.datetime):
            data_linha = data_linha.date()
        if data_linha is None:
            continue
        if data_inicio <= data_linha <= data_fim:
            linhas_candidatas.append(r)
        elif data_linha > data_fim and primeira_linha_apos_periodo is None:
            primeira_linha_apos_periodo = r

    if linhas_candidatas:
        bloco_continuo = linhas_candidatas == list(
            range(linhas_candidatas[0], linhas_candidatas[-1] + 1)
        )
        if not bloco_continuo:
            raise RuntimeError(
                "As linhas dentro do periodo do extrato nao formam um bloco "
                f"continuo ({linhas_candidatas[0]}-{linhas_candidatas[-1]}). "
                "Revise a planilha manualmente antes de rodar o script de novo."
            )
        bloco_inicio = linhas_candidatas[0]
        bloco_fim = linhas_candidatas[-1]
    else:
        bloco_inicio = primeira_linha_apos_periodo or (ultima_linha_planilha + 1)
        bloco_fim = None

    # Captura (snapshot) as linhas existentes do bloco, agrupadas por
    # data, antes de remover qualquer coisa.
    sheet_por_dia = defaultdict(list)
    if bloco_fim is not None:
        for r in range(bloco_inicio, bloco_fim + 1):
            data_linha = ws.cell(row=r, column=COL_DATA).value
            if isinstance(data_linha, datetime.datetime):
                data_linha = data_linha.date()
            credito = ws.cell(row=r, column=COL_CREDITO).value
            debito = ws.cell(row=r, column=COL_DEBITO).value
            valor = credito if credito else -(debito or 0)
            snapshot = {
                "data": data_linha,
                "lancamento": ws.cell(row=r, column=COL_LANCAMENTO).value,
                "descricao": ws.cell(row=r, column=COL_DESCRICAO).value,
                "credito": credito,
                "debito": debito,
                "categoria": ws.cell(row=r, column=COL_CATEGORIA).value,
                "valor": round(float(valor), 2),
                "estilos": {c: copy(ws.cell(row=r, column=c)._style) for c in range(1, 8)},
            }
            sheet_por_dia[data_linha].append(snapshot)

    # Decide, dia a dia, se mantem o que ja esta na planilha ou substitui
    # pelos lancamentos reais do extrato.
    todas_datas = sorted(set(extrato_por_dia) | set(sheet_por_dia))
    plano = []  # lista de ("manter", snapshot) ou ("novo", (data,texto,valor))
    dias_mantidos = 0
    dias_substituidos = 0
    preenchidos_retroativamente = 0
    for data in todas_datas:
        valores_extrato = sorted(round(v, 2) for _, v in extrato_por_dia.get(data, []))
        valores_planilha = sorted(s["valor"] for s in sheet_por_dia.get(data, []))
        if valores_extrato == valores_planilha and valores_extrato:
            dias_mantidos += 1
            # Preenche LANCAMENTO em branco cruzando por valor.
            textos_por_valor = defaultdict(list)
            for texto, v in extrato_por_dia[data]:
                textos_por_valor[round(v, 2)].append(texto)
            snaps_por_valor = defaultdict(list)
            for s in sheet_por_dia[data]:
                snaps_por_valor[s["valor"]].append(s)
            for v, textos in textos_por_valor.items():
                for texto, s in zip(textos, snaps_por_valor[v]):
                    if not s["lancamento"]:
                        s["lancamento"] = texto
            # Preenchimento retroativo: se DESCRICAO e/ou CATEGORIA
            # estiverem em branco numa linha ja existente, mas o mesmo
            # texto de LANCAMENTO ja foi classificado em outra linha da
            # planilha, reaplica essa classificacao aqui tambem.
            for s in sheet_por_dia[data]:
                if not s["lancamento"]:
                    continue
                descricao_auto, categoria_auto = mapa_desc_cat.get(s["lancamento"], (None, None))
                preencheu = False
                if not s["descricao"] and descricao_auto:
                    s["descricao"] = descricao_auto
                    preencheu = True
                if not s["categoria"] and categoria_auto:
                    s["categoria"] = categoria_auto
                    preencheu = True
                if preencheu:
                    preenchidos_retroativamente += 1
            for s in sheet_por_dia[data]:
                plano.append(("manter", s))
        else:
            dias_substituidos += 1
            for texto, v in extrato_por_dia.get(data, []):
                plano.append(("novo", (data, texto, v)))

    if bloco_fim is not None:
        ws.delete_rows(bloco_inicio, amount=(bloco_fim - bloco_inicio + 1))

    ws.insert_rows(bloco_inicio, amount=len(plano))

    template_row = max(2, bloco_inicio - 1)
    preenchidos_automaticamente = 0
    linhas_novas = 0

    for i, (tipo, item) in enumerate(plano):
        linha_atual = bloco_inicio + i

        if tipo == "manter":
            s = item
            ws.cell(row=linha_atual, column=COL_DATA, value=datetime.datetime(s["data"].year, s["data"].month, s["data"].day))
            ws.cell(row=linha_atual, column=COL_LANCAMENTO, value=s["lancamento"])
            ws.cell(row=linha_atual, column=COL_DESCRICAO, value=s["descricao"])
            ws.cell(row=linha_atual, column=COL_CREDITO, value=s["credito"])
            ws.cell(row=linha_atual, column=COL_DEBITO, value=s["debito"])
            ws.cell(row=linha_atual, column=COL_CATEGORIA, value=s["categoria"])
            for c in range(1, 8):
                ws.cell(row=linha_atual, column=c)._style = s["estilos"][c]
        else:
            data, texto, valor = item
            linhas_novas += 1
            descricao_auto, categoria_auto = mapa_desc_cat.get(texto, (None, None))
            if descricao_auto or categoria_auto:
                preenchidos_automaticamente += 1
            ws.cell(row=linha_atual, column=COL_DATA, value=datetime.datetime(data.year, data.month, data.day))
            ws.cell(row=linha_atual, column=COL_LANCAMENTO, value=texto)
            ws.cell(row=linha_atual, column=COL_DESCRICAO, value=descricao_auto)
            ws.cell(row=linha_atual, column=COL_CREDITO, value=valor if valor > 0 else None)
            ws.cell(row=linha_atual, column=COL_DEBITO, value=abs(valor) if valor < 0 else None)
            ws.cell(row=linha_atual, column=COL_CATEGORIA, value=categoria_auto)
            for c in range(1, 8):
                ws.cell(row=linha_atual, column=c)._style = copy(ws.cell(row=template_row, column=c)._style)

    # A cadeia de formulas de SALDO e regerada do zero, pois delete_rows/
    # insert_rows nao reescrevem as referencias de linha dentro do texto
    # das formulas ja existentes.
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=COL_SALDO, value=f"=F{r - 1}+D{r}-E{r}")

    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 15.75

    wb.save(XLSX_PATH)
    print(f"Dias ja reconciliados e mantidos: {dias_mantidos} "
          f"({preenchidos_retroativamente} linha(s) com DESCRICAO/CATEGORIA preenchida(s) retroativamente)")
    print(f"Dias substituidos pelo extrato: {dias_substituidos} ({linhas_novas} lancamento(s) novo(s), "
          f"sendo {preenchidos_automaticamente} com DESCRICAO/CATEGORIA preenchidas automaticamente)")
    print(f"Planilha salva em {XLSX_PATH}")


if __name__ == "__main__":
    main()
